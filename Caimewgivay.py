from tkinter import *
from tkinter import ttk, filedialog, messagebox
import threading, time, random, os, requests
import webbrowser

from seleniumbase import Driver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select

from openpyxl import Workbook, load_workbook


# ================= CONFIG =================
AUTO_SAVE_FILE = "auto_save.xlsx"

TIME_DELAY_MIN = 1
TIME_DELAY_MAX = 3

start_time_tab1 = 0

CHROME_W = 500
CHROME_H = 400

# reset chrome sau X request
RESET_CHROME_EVERY = 50

BASE_URL = "https://masothue.com/"


# ================= GLOBAL =================
input_list_tab1 = []

done_tab1 = 0
total_tab1 = 0

current_index_tab1 = 0

is_running = False


# ================= KEY ======================
SHEET_CSV = "https://docs.google.com/spreadsheets/d/1Xs81jas5b0DCdk5iVWdczsJJTlqfh1lsR4o9Iwn9OHg/export?format=csv"


def check_key_online(user_key):
    try:
        res = requests.get(SHEET_CSV, timeout=10)

        if res.status_code != 200:
            return False

        data = res.text.splitlines()

        for line in data:
            cols = line.split(",")

            key = cols[0].strip()
            status = cols[1].strip() if len(cols) > 1 else "0"

            if key == user_key and status == "1":
                return True

        return False

    except Exception as e:
        print("Lỗi check key:", e)
        return False


# ================= PROGRESS =================
def update_progress_tab1():
    if total_tab1 == 0:
        return

    percent = (done_tab1 / total_tab1) * 100
    progress_var1.set(percent)

    elapsed = time.time() - start_time_tab1

    if done_tab1 > 0:
        rate = done_tab1 / elapsed
        remaining = (total_tab1 - done_tab1) / rate
        eta = time.strftime("%H:%M:%S", time.gmtime(remaining))
    else:
        rate = 0
        eta = "--:--:--"

    label_percent1.config(
        text=f"{percent:.0f}% | {done_tab1}/{total_tab1} | ETA: {eta} | {rate:.1f} req/s"
    )


# ================= SAVE =================
def save_realtime(row):
    if os.path.exists(AUTO_SAVE_FILE):
        wb = load_workbook(AUTO_SAVE_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Input", "MST", "URL"])

    ws.append([
        row[0],
        f"'{row[1]}",  # giữ số 0 đầu
        row[2]
    ])

    wb.save(AUTO_SAVE_FILE)


# ================= EXPORT =================
def export_excel(tree):
    if not tree.get_children():
        messagebox.showwarning("Lỗi", "Không có dữ liệu")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )

    if not file_path:
        return

    wb = Workbook()
    ws = wb.active

    ws.append(["Input", "MST", "URL"])

    for item in tree.get_children():
        values = tree.item(item)["values"]

        ws.append([
            values[0],
            f"'{values[1]}",  # giữ số 0 đầu
            values[2]
        ])

    wb.save(file_path)

    messagebox.showinfo("OK", "Export thành công")


# ================= LOAD EXCEL =================
def load_excel(tab):
    global input_list_tab1
    global current_index_tab1

    current_index_tab1 = 0

    file = filedialog.askopenfilename(
        filetypes=[("Excel", "*.xlsx")]
    )

    if not file:
        return

    wb = load_workbook(file)
    ws = wb.active

    data = []

    for r in ws.iter_rows(min_row=1, max_col=1):
        v = r[0].value

        if v:
            data.append(str(v))

    if tab == 1:
        input_list_tab1 = data

        tree1.delete(*tree1.get_children())

        for item in data:
            tree1.insert("", "end", values=(item, "", ""))

    messagebox.showinfo("OK", f"Load {len(data)} dòng")


# ================= SETTINGS =================
def apply_settings():
    global TIME_DELAY_MIN
    global TIME_DELAY_MAX
    global CHROME_W
    global CHROME_H
    global RESET_CHROME_EVERY

    TIME_DELAY_MIN = int(entry_min.get())
    TIME_DELAY_MAX = int(entry_max.get())

    CHROME_W = int(entry_w.get())
    CHROME_H = int(entry_h.get())

    RESET_CHROME_EVERY = int(entry_reset.get())

    messagebox.showinfo("OK", "Đã cập nhật setting")


# ================= DRIVER =================
def create_driver(proxy=None):
    options = {
        "uc": True,
        "headless": False,
        "agent": None
    }

    if proxy:
        options["proxy"] = proxy

    driver = Driver(**options)

    driver.set_window_size(CHROME_W, CHROME_H)
    driver.set_window_position(0, 0)

    return driver


def extract_mst(url):
    try:
        return url.split("/")[-1].split("-")[0]
    except:
        return ""


# ================= TREE UPDATE =================
def update_tree(tree, i, inp, mst, url):
    def _u():
        global done_tab1

        items = tree.get_children()

        if i < len(items):
            tree.item(items[i], values=(inp, mst, url))
        else:
            tree.insert("", "end", values=(inp, mst, url))

        if tree == tree1:
            done_tab1 += 1
            update_progress_tab1()

    root.after(0, _u)


# ================= VALID =================
def is_valid_mst(mst, url):
    if not mst or not mst.isdigit():
        return False

    if "Search" in url or "?q=" in url:
        return False

    if url.strip() == BASE_URL:
        return False

    return True


# ================= WORKER =================
def worker_tab1():
    global current_index_tab1

    driver = create_driver()

    request_count = 0

    for i in range(current_index_tab1, len(input_list_tab1)):

        if not is_running:
            current_index_tab1 = i
            break

        raw = input_list_tab1[i]

        attempt = 0
        success = False

        while attempt < 2:
            try:
                driver.get(BASE_URL)

                try:
                    select_box = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located(
                            (By.ID, "product_cat")
                        )
                    )

                    Select(select_box).select_by_value("personalTax")

                except:
                    pass

                search = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.ID, "search")
                    )
                )

                search.clear()
                search.send_keys(raw)
                search.send_keys(Keys.ENTER)

                time.sleep(
                    random.uniform(
                        TIME_DELAY_MIN,
                        TIME_DELAY_MAX
                    )
                )

                url = driver.get_current_url()
                mst = extract_mst(url)

                if is_valid_mst(mst, url):

                    update_tree(
                        tree1,
                        i,
                        raw,
                        mst,
                        url
                    )

                    save_realtime([
                        raw,
                        mst,
                        url
                    ])

                    success = True
                    break

                else:
                    attempt += 1
                    time.sleep(1)

            except Exception as e:
                print("Retry lỗi:", e)

                attempt += 1
                time.sleep(1)

        if not success:
            update_tree(
                tree1,
                i,
                raw,
                "0",
                ""
            )

        current_index_tab1 = i + 1

        # ================= RESET CHROME =================
        request_count += 1

        if request_count >= RESET_CHROME_EVERY:
            try:
                driver.quit()
            except:
                pass

            print("Đang reset Chrome...")

            time.sleep(2)

            driver = create_driver()

            request_count = 0

            print("Đã reset Chrome")

    try:
        driver.quit()
    except:
        pass


# ================= ETA =================
def format_time(sec):
    sec = int(sec)

    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60

    return f"{h:02}:{m:02}:{s:02}"


def update_eta_loop():
    while is_running:
        try:
            if done_tab1 > 0 and start_time_tab1:

                elapsed = time.time() - start_time_tab1

                speed = done_tab1 / elapsed

                remaining = total_tab1 - done_tab1

                eta = remaining / speed if speed > 0 else 0

                percent = (
                    done_tab1 / total_tab1
                ) * 100

                text = (
                    f"{percent:.0f}% | "
                    f"{done_tab1}/{total_tab1} | "
                    f"ETA: {format_time(eta)} | "
                    f"{speed:.2f} req/s"
                )

                root.after(
                    0,
                    lambda: label_percent1.config(
                        text=text
                    )
                )

        except:
            pass

        time.sleep(1)


# ================= START =================
def start_tab1():
    global is_running
    global total_tab1
    global done_tab1
    global start_time_tab1

    user_key = entry_key.get().strip()

    print(user_key)

    if not check_key_online(user_key):
        messagebox.showerror(
            "Key",
            "Key không hợp lệ hoặc hết hạn"
        )
        return

    if not input_list_tab1:
        messagebox.showwarning(
            "Lỗi",
            "Chưa load Excel"
        )
        return

    is_running = True

    total_tab1 = len(input_list_tab1)
    done_tab1 = 0

    start_time_tab1 = time.time()

    progress_var1.set(0)

    threading.Thread(
        target=worker_tab1,
        daemon=True
    ).start()

    threading.Thread(
        target=update_eta_loop,
        daemon=True
    ).start()


def stop_tool():
    global is_running
    is_running = False


# ================= UI =================
root = Tk()

root.title("Vì Yêu Cứ Đâm Dow")

root.geometry("900x500")
root.minsize(900, 500)

notebook = ttk.Notebook(root)
notebook.pack(fill="both", expand=True)


# ================= TAB 1 =================
tab1 = Frame(notebook)

notebook.add(tab1, text="Worker 1")

frame_btn1 = Frame(tab1)
frame_btn1.pack(fill="x")

Button(
    frame_btn1,
    text="Load Excel",
    command=lambda: load_excel(1)
).pack(side="left", padx=5)

Button(
    frame_btn1,
    text="Start",
    command=start_tab1
).pack(side="left", padx=5)

Button(
    frame_btn1,
    text="Stop",
    command=stop_tool
).pack(side="left", padx=5)

Button(
    frame_btn1,
    text="Export",
    command=lambda: export_excel(tree1)
).pack(side="left", padx=5)


frame_tree1 = Frame(tab1)
frame_tree1.pack(fill="both", expand=True)

tree1 = ttk.Treeview(
    frame_tree1,
    columns=("Input", "MST", "URL"),
    show="headings"
)

for c in ("Input", "MST", "URL"):
    tree1.heading(c, text=c)
    tree1.column(c, width=200)

scroll_y1 = Scrollbar(
    frame_tree1,
    orient="vertical",
    command=tree1.yview
)

tree1.configure(
    yscrollcommand=scroll_y1.set
)

tree1.pack(
    side="left",
    fill="both",
    expand=True
)

scroll_y1.pack(
    side="right",
    fill="y"
)


progress_var1 = DoubleVar()

ttk.Progressbar(
    tab1,
    variable=progress_var1,
    maximum=100
).pack(fill="x")

label_percent1 = Label(
    tab1,
    text="0%"
)

label_percent1.pack()


# ================= TAB 3 =================
tab3 = Frame(notebook)

notebook.add(tab3, text="Setting")

frame_main = Frame(
    tab3,
    padx=10,
    pady=10
)

frame_main.pack(
    fill="both",
    expand=True
)

frame_main.columnconfigure(
    1,
    weight=1
)


# ===== DELAY MIN =====
Label(
    frame_main,
    text="Delay Min"
).grid(
    row=0,
    column=0,
    sticky="w",
    pady=5
)

entry_min = Entry(frame_main)
entry_min.insert(0, "2")

entry_min.grid(
    row=0,
    column=1,
    padx=5,
    sticky="ew"
)


# ===== DELAY MAX =====
Label(
    frame_main,
    text="Delay Max"
).grid(
    row=1,
    column=0,
    sticky="w",
    pady=5
)

entry_max = Entry(frame_main)
entry_max.insert(0, "4")

entry_max.grid(
    row=1,
    column=1,
    padx=5,
    sticky="ew"
)


# ===== WIDTH =====
Label(
    frame_main,
    text="Chrome Width"
).grid(
    row=2,
    column=0,
    sticky="w",
    pady=5
)

entry_w = Entry(frame_main)
entry_w.insert(0, "500")

entry_w.grid(
    row=2,
    column=1,
    padx=5,
    sticky="ew"
)


# ===== HEIGHT =====
Label(
    frame_main,
    text="Chrome Height"
).grid(
    row=3,
    column=0,
    sticky="w",
    pady=5
)

entry_h = Entry(frame_main)
entry_h.insert(0, "400")

entry_h.grid(
    row=3,
    column=1,
    padx=5,
    sticky="ew"
)


# ===== RESET CHROME =====
Label(
    frame_main,
    text="Reset Chrome Every"
).grid(
    row=4,
    column=0,
    sticky="w",
    pady=5
)

entry_reset = Entry(frame_main)
entry_reset.insert(0, "50")

entry_reset.grid(
    row=4,
    column=1,
    padx=5,
    sticky="ew"
)


# ===== KEY =====
def toggle_key():
    if entry_key.cget("show") == "*":
        entry_key.config(show="")
    else:
        entry_key.config(show="*")


Label(
    frame_main,
    text="License"
).grid(
    row=6,
    column=0,
    sticky="w",
    pady=5
)

entry_key = Entry(
    frame_main,
    show="*"
)

entry_key.grid(
    row=6,
    column=1,
    sticky="ew",
    padx=5
)

Button(
    frame_main,
    text="Show Key",
    command=toggle_key
).grid(
    row=7,
    column=1
)


# ===== FACEBOOK =====
def open_url(event):
    webbrowser.open(
        "https://www.facebook.com/tran.long.vu.406156/"
    )


label_url = Label(
    frame_main,
    text="Không Gì Tốt Hơn Bằng Sự Thương Cảm Và Sự Hỗ Trợ, Vậy Nên Hãy Cứu Tôi Khỏi Sự Nghèo Khổ",
    fg="blue",
    cursor="hand2",
    font=("Segoe UI", 10, "underline")
)

label_url.grid(
    row=5,
    column=0,
    columnspan=2,
    sticky="w",
    pady=5
)

label_url.bind(
    "<Button-1>",
    open_url
)


# ===== APPLY =====
frame_bottom = Frame(
    tab3,
    pady=5
)

frame_bottom.pack(
    side="bottom",
    fill="x"
)

Button(
    frame_bottom,
    text="Apply Settings",
    command=apply_settings,
    height=1
).pack(
    fill="x",
    padx=4
)


root.mainloop()